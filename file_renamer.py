# file_renamer: file_renamer.py

from os import path, rename
from openpyxl import load_workbook

def print_rev_to_console(name, rev, rev_date):
    """Print a revision block to console."""
    print(f"{name} by S. Cassier.\nRevision: {rev} - {rev_date}.\n")

def read_excel():
    filepath = 'c:/local/file_renamer.xlsx'
    workbook = load_workbook(filename=filepath, read_only=True)
    worksheet = workbook.active
    
    for line in worksheet.iter_rows(min_row=2, min_col=1):
        yield [cell.value for cell in line]

def rename_files():
    input_path = input("Enter the path containing the files:\n> ")
    list_files_to_rename = read_excel()

    for line in list_files_to_rename:
        if input_path.endswith("/"):
            old_name = input_path + line[0]
            new_name = input_path + line[1]
        else:
            old_name = input_path + "/" + line[0]
            new_name = input_path + "/" + line[1]
        
        if verify_path(old_name):
            rename(old_name, new_name)
        else:
            print(f"{old_name} not found.")
    
    print("Done.")
    input("Press enter to exit.")

def verify_path(filepath):
    return True if path.isfile(filepath) else False

if __name__ == "__main__":
    print_rev_to_console('file_renamer', 0.1, '2020-07-06')
    rename_files()