# file_renamer: _c_filerenamerui.py
# Import statement(s).
from os import path, rename
from sys import exit
from tkinter import Button, E, Entry, filedialog, Frame, StringVar, W
from openpyxl import load_workbook

class FileRenamerUI(Frame):
    """Definition for class FileRenamerUI."""

    def __init__(self, parent, rev, rev_date):
        """Class constructor."""
        # Inheritance from parent-object is required: tkinter.Tk() is called
        # in the application using class FileRenamerUI.
        super(FileRenamerUI, self).__init__(parent)
        # Call functions to build and maintain the UI-elements.
        self.gui()
        self.pack()

    def _yield_from_xlsx_(self, filepath):
        """Read .xlsx-file containing the renaming-matrix and yield each line
        in a list-form.
        """
        self.wb = load_workbook(filename=filepath, read_only=True)
        self.ws = self.wb.active

        for line in self.ws.iter_rows(min_row=2, min_col=1):
            yield [cell.value for cell in line]

    def gui(self):
        """GUI definition."""
        # Variable(s).
        self.path_files = StringVar()
        self.path_xlsx = StringVar()
        # Create the widget(s).
        self.btn_files = Button(self, text="Select folder", anchor=W, 
            command=self.cmd_select_files)
        self.btn_quit = Button(self, text="Quit", command=self.cmd_quit)
        self.btn_rename = Button(self, text="Start rename", 
            command=self.cmd_rename)
        self.btn_xlsx = Button(self, text="Select xlsx", anchor=W, 
            command=self.cmd_select_xlsx)
        self.ent_files = Entry(self, text=self.path_files, width=70)
        self.ent_xlsx = Entry(self, text=self.path_xlsx, width=70)
        # Attach the widget(s).
        self.btn_files.grid(row=0, column=0, padx=2, pady=2, sticky=W+E)
        self.btn_quit.grid(row=2, column=3, columnspan=2, padx=2, pady=2,
            sticky=W+E)
        self.btn_rename.grid(row=2, column=1, columnspan=2, padx=2, pady=2,
            sticky=W+E)
        self.btn_xlsx.grid(row=1, column=0, padx=2, pady=2, sticky=W+E)
        self.ent_files.grid(row=0, column=1, columnspan=4, ipady=2, padx=2, pady=2)
        self.ent_xlsx.grid(row=1, column=1, columnspan=4, ipady=2, padx=2, pady=2)

    def cmd_quit(self):
        """Close the class and the application calling the class."""
        exit(0)

    def cmd_rename(self):
        """Rename the files."""
        # Variable(s).
        self.path_files = StringVar()
        self.path_xlsx = StringVar()
        self.path_files = self.ent_files.get()
        self.path_xlsx = self.ent_xlsx.get()
        # Process.
        for line in self._yield_from_xlsx_(self.path_xlsx):
            old_name = self.path_files + "/" + line[0]
            new_name = self.path_files + "/" + line[1]

            if path.isfile(old_name):
                rename(old_name, new_name)

    def cmd_select_files(self):
        """Select the folder with the files to be renamed."""
        # Variable(s).
        self.path_files = StringVar()
        input_path = filedialog.askdirectory(initialdir='c:/local/',
            title="Select the folder with the files to rename")
        self.path_files.set(input_path)
        # Redraw and update the relevant widget.
        self.ent_files.config(text=self.path_files)

    def cmd_select_xlsx(self):
        """Select the .xlsx-file with the conversion-table for the renaming
        action.
        """
        # Variable(s).
        self.path_xlsx = StringVar()
        input_path = filedialog.askopenfilename(initialdir='c:/local/',
            title="Select the conversion-matrix",
            filetype=[('XLSX files', '*.xlsx')])
        self.path_xlsx.set(input_path)
        # Redraw and update the relevant widget.
        self.ent_xlsx.config(text=self.path_xlsx)
